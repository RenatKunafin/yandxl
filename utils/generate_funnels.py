import sys
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import Font, Fill, PatternFill


prefix = 'Открытие брокерского счета и ИИС'
source_prexixes = ['', 'ln_iis1_', 'ln_iis2_']
# sources = [
#     'dashboard',
#     'marketplace',
#     'marketplace_invest',
#     'search_brok',
#     'search_iis',
#     'banner',
#     'webiis',
#     'webdios',
#     'webcurrency',
#     'webcurrencyoffer',
#     'webonboarding',
#     'webbonds',
#     'webbondslist',
#     'webbondsiis',
#     'webbondsoffer',
#     'webstocks',
#     'webstocksoffer',
#     'webstocksfaq',
#     'webforeignstocks',
#     'webforeignstockstax',
#     'webforeignstocksoffer',
#     'webetf',
#     'webetfoffer',
#     'webofzn',
#     'webeurobonds',
#     'websnotes',
#     'webotc',
#     'webotcrates',
#     'webforts',
#     'webfortsoffer',
#     'webinvbrokmain',
#     'websbinvestor',
#     'websbinvestoridea',
#     'webquik',
#     'websposoby',
#     'webmargin',
#     'webrepo',
#     'webqual',
#     'webhowtostart',
#     'websmartinvest'
#     ]
sources = [
    'marketplace',
    'marketplace_invest',
    'search_brok',
    'search_iis',
    'all_services'
    ]
steps = [
        '[SBOL_WEB_PRO][brokerage.management]{}.Налоговая информация.Открытие',
        '[SBOL_WEB_PRO][brokerage.management]{}.Параметры счёта.Открытие',
        '[SBOL_WEB_PRO][brokerage.management]{}.Контакты.Один телефон.Открытие',
        '[SBOL_WEB_PRO][brokerage.management]{}.Контакты.Несколько телефонов.Открытие',
        '[SBOL_WEB_PRO][brokerage.management]{}.Вывод средств.Открытие',
        '[SBOL_WEB_PRO][brokerage.management]{}.ИИС.Открытие',
        '[SBOL_WEB_PRO][brokerage.management]{}.Подтверждение.Открытие',
        '[SBOL_WEB_PRO][brokerage.management]{}.Статусный экран.Открытие'
    ]

def getAllStepsForSource (source):
    values = []
    for step in steps:
        for s_prefix in source_prexixes:
            values.append(prefix+'.'+s_prefix+source+'.'+step)
    print('!>', values)
    return values

def getStepForAllSources (step):
    values = []
    for source in sources:
        for s_prefix in source_prexixes:
            # values.append(prefix+'.'+s_prefix+source+'.'+step)
            values.append(prefix+'.'+s_prefix+source+'.'+step)
    values.append('next')
    print('!>', values)
    return values

def generateFunnelByStep ():
    wb = load_workbook('./Brokerage_ymetrics_report.xlsx')
    ws_dashboard = wb['Дашборд']
    ws_funnel = wb['СБОЛПРО-ОДБО']
    variants = []
    done = False
    prev_variant = ''
    for step in steps:
        variants += getStepForAllSources(step)

    for variant in variants:
        done = False
        if (variant == 'next'):
            
            temp = prev_variant.split('.')
            temp.pop(1)
            ws_funnel.append(['.'.join(temp), ''])
            ws_funnel.append(['', ''])
            continue
        for row in ws_dashboard.iter_rows(min_row=ws_dashboard.min_row+1, max_row=ws_dashboard.max_row, min_col=1, max_col=1):
            if (done is False):
                for cell in row:
                    if (cell.value == variant):
                        usersValue = cell.offset(row=0, column=+1)
                        # print('!>>', cell.value, usersValue.value)
                        ws_funnel.append([cell.value, usersValue.value])
                        done = True
                    elif (done is True):
                        break
                    else:
                        continue
            else:
                break
        prev_variant = variant
        if (done is False):
            ws_funnel.append([variant, 0])
        

    wb.save('./Brokerage_ymetrics_report.xlsx')


# Сбор воронки по источникам - шаг1, ln1_шаг1, ln2_шаг1, шаг2, ln1_шаг2, ln2_шаг2
def generateFunnelBySourse ():
    wb = load_workbook('./Brokerage_ymetrics_report.xlsx')
    ws_dashboard = wb['Дашборд']
    ws_funnel = wb['ОДБО']
    source_variants = []
    for source in sources:
        source_variants += getAllStepsForSource(source)
    for variant in source_variants:
        for row in ws_dashboard.iter_rows(min_row=ws_dashboard.min_row+1, max_row=ws_dashboard.max_row, min_col=1, max_col=1):
            for cell in row:
                if (cell.value == variant):
                    usersValue = cell.offset(row=0, column=+1)
                    # print('!>>', cell.value, usersValue.value)
                    ws_funnel.append([cell.value, usersValue.value])
                else:
                    continue

    wb.save('./Brokerage_ymetrics_report.xlsx')
                

def main(argv):
    generateFunnelByStep()
    
if __name__ == "__main__":
    main(sys.argv[1:])