from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import Font, Fill, PatternFill
from datetime import datetime
from datetime import timedelta
from hashlib import md5
import time
import json

from openpyxl.workbook.defined_name import ROW_RANGE_RE


class Excel2:
    def __init__(self, cfg, data):
        self.data = data
        self.query = data['query']
        self.path_to_wb = cfg.get('smtp', 'PATH') + cfg.get('excel', 'WB_NAME') + '_' + str(time.time()) + '.xlsx'
        self.dashboard_ws_name = cfg.get('excel', 'DASHBOARD_WS_NAME')
        self.odbo_ws_name = cfg.get('excel', 'ODBO_FUNNEL_SHEET_NAME')
        self.odbo_titles = cfg.get('excel', 'ODBO_FUNNEL_TITLES').split(',')
        self.odbo_steps = cfg.get('excel', 'ODBO_FUNNEL_STEPS').split(',')
        self.titles_dashboard = cfg.get('excel', 'ROW_TITLES_DASHBOARD2').split(',')
        self.titles = cfg.get('excel', 'ROW_TITLES2').split(',')
        self.max_ws_name_length = int(cfg.get('excel', 'MAX_WS_NAME_LENGTH'))

    def _build_dashboard_row(self, dimensions, metrics):
        entry = []
        for d in dimensions:
            entry.append(d['name'])
        return entry + metrics

    def _build_dashboard(self, wb):
        ws_dashboard = wb.active
        ws_dashboard.delete_cols(1, 10)
        ws_dashboard.title = self.dashboard_ws_name
        ws_dashboard.append(self.titles_dashboard)
        entries = []
        for d in self.data['data']:
            entries.append(self._build_dashboard_row(d['dimensions'], d['metrics']))
        for row in entries:
            ws_dashboard.append(row)

    def _dataToList(self, dimensions, metrics):
        # result = []
        # for d in dimensions:
        #     if (d['name'] is not None):
        #         result.append(d['name'])
        return [dimensions[4]['name'], dimensions[0]['name'], dimensions[3]['name'], metrics[0]]

    def _build_odbo_funnel(self, wb):
        ws_odbo = wb.create_sheet(self.odbo_ws_name)
        ws_odbo.title = self.odbo_ws_name
        ws_odbo.append(self.odbo_titles)
        tree = { i : {} for i in self.odbo_steps }
        for d in self.data['data']:
            if (d['dimensions'][2]['name'] == 'Открытие брокерского счета'):
                step = d['dimensions'][4]['name']
                source = d['dimensions'][3]['name']
                # print('!>', source, tree)
                if source in tree[step]:
                    if d['dimensions'][5]['name'] == 'Открытие':
                        tree[step][source].append(self._dataToList(d['dimensions'], d['metrics']))
                    elif d['dimensions'][4]['name'] == 'Контакты' and (d['dimensions'][5]['name'] == 'Один телефон' or d['dimensions'][5]['name'] == 'Несколько телефонов'):
                        for i in range(len(tree[step][source])):
                            if tree[step][source][i][0] == 'Контакты':
                                tree[step][source][i][-1] = tree[step][source][i][-1] + d['metrics'][0]
                elif d['dimensions'][4]['name'] == 'Контакты' and d['dimensions'][5]['name'] == 'Продолжить_клик':
                    continue
                else:
                    tree[step][source] = [self._dataToList(d['dimensions'], d['metrics'])]
        f = open("tree.json", "w")
        f.write(json.dumps(tree))
        f.close()
        return tree
        # for branch in tree:
        #     for source in tree[branch]:
        #         for row in tree[branch][source]:
        #             # print('!!!', tree[branch][source])
        #             ws_odbo.append(row)

    def _update_funnel(self, wb, data):


    def init_wb(self):
        wb = Workbook()
        self._build_dashboard(wb)
        odbo_funnel = self._build_odbo_funnel(wb)
        self._update_funnel(wb, odbo_funnel)
        wb.save(self.path_to_wb)
        print('excel ready')
