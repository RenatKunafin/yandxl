import sys
from hashlib import md5
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import Font, Fill, PatternFill

def clean ():
    items = [
        'brokerage.management.application.metric.ODBO.ln1_marketplace_invest.brokerage.management.application.metric.confirmation.Открытие',
        'brokerage.management.application.metric.ODBO.ln1_marketplace_invest.brokerage.management.application.metric.confirmation.Подтвердить_клик',
        'brokerage.management.application.metric.ODBO.ln1_marketplace_invest.brokerage.management.application.metric.status.Открытие',
        'brokerage.management.application.metric.ODBO.ln1_marketplace_invest.Вывод средств.Открытие',
        'brokerage.management.application.metric.ODBO.ln1_marketplace_invest.Вывод средств.Продолжить_клик',
        'brokerage.management.application.metric.ODBO.ln1_marketplace_invest.ИИС.Открытие',
        'brokerage.management.application.metric.ODBO.ln1_marketplace_invest.ИИС.Продолжить_клик',
        'brokerage.management.application.metric.ODBO.ln1_marketplace_invest.Контакты.brokerage.management.application.metric.phones.single.Открытие',
        'brokerage.management.application.metric.ODBO.ln1_marketplace_invest.Контакты.Продолжить_клик',
        'brokerage.management.application.metric.ODBO.ln1_marketplace_invest.Налоговая информация.Открытие',
        'brokerage.management.application.metric.ODBO.ln1_marketplace_invest.Налоговая информация.Продолжить_клик',
        'brokerage.management.application.metric.ODBO.ln1_marketplace_invest.Параметры счёта.Открытие',
        'brokerage.management.application.metric.ODBO.ln1_marketplace_invest.Параметры счёта.Продолжить_клик',
        'Актив.Внебиржевой рынок.Количество.График',
        'Актив.Внебиржевой рынок.Рыночная цена.График',
        'Актив.Внебиржевой рынок.Стоимость.График',
        'Актив.Срочный рынок.Вариационная маржа.График',
        'Актив.Срочный рынок.Количество.График',
        'Актив.Срочный рынок.Рыночная цена.График',
        'Актив.Фондовый рынок.Количество.График',
        'Актив.Фондовый рынок.Рыночная цена.График',
        'Актив.Фондовый рынок.Стоимость.График',
        'Мои рынки.Все рынки.Активы.Актив Фондового рынка_Клик',
        'Мои рынки.Все рынки.Активы.Включение переключателя у любого актива_Клик',
        'Мои рынки.Все рынки.Активы.Выключение переключателя у любого актива_Клик',
        'Мои рынки.Все рынки.Активы.Рубли_Клик',
        'Список счетов.Счета развернуты.Отчеты.2НДФЛ',
        'Список счетов.Счета развернуты.Отчеты.Брокерский отчет',
        'Заявление на открытие ДБО.Успешное подтверждение по СМС',
        'Заявление на открытие ДБО.Шаг_1 Выбор рынка.Открытие шага',
        'Заявление на открытие ДБО.Шаг_10 Указание адреса постоянной регистрации.Открытие шага',
        'Заявление на открытие ДБО.Шаг_11 Подтверждение налогового резидентства только РФ.Открытие шага',
        'Заявление на открытие ДБО.Шаг_12 Ввод мобильного телефона.Открытие шага',
        'Заявление на открытие ДБО.Шаг_13 Ввод email адреса.Открытие шага',
        'Заявление на открытие ДБО.Шаг_14 подтверждение.Кнопка_Отправить_заявление_Клик',
        'Заявление на открытие ДБО.Шаг_2 Выбор тарифа.Открытие шага',
        'Заявление на открытие ДБО.Шаг_3 Указание счетов вывода.Открытие шага',
        'Заявление на открытие ДБО.Шаг_4 Возможность получения доп. дохода.Открытие шага',
        'Заявление на открытие ДБО.Шаг_5 Использование заемных средств.Открытие шага',
        'Заявление на открытие ДБО.Шаг_6 ИИС.Открытие шага',
        'Заявление на открытие ДБО.Шаг_7 Цель открытия счета.Открытие шага',
        'Заявление на открытие ДБО.Шаг_8 Страна рождения.Открытие шага',
        'Заявление на открытие ДБО.Шаг_9 Подтверждение персональных данных.Открытие шага',
        'Заявление на открытие ИИС.Успешное подтверждение по СМС',
        'Заявление на открытие ИИС.Шаг 4 Подтверждение.Кнопка_Отправить_заявление_Клик',
        'Заявление на открытие ИИС.Шаг_1 Выбор основного договора.Открытие шага',
        'Заявление на открытие ИИС.Шаг_2 Выбор рынка.Открытие шага',
        'Заявление на открытие ИИС.Шаг_3 Информация о тарифе.Открытие шага'
    ]
    wb = load_workbook('./Brokerage_ymetrics_report.xlsx')
    ws_dashboard = wb['Дашборд']
    found = 0
    removed = 0
    print('BEFORE', len(wb.sheetnames))
    for item in items:
        name = str(md5(item.encode('UTF-8')).hexdigest()[:-1])
        try:
            wb[name]
            found += 1
            wb.remove(wb[name])
            removed += 1
            print('FOUND', item, name)
            for row in ws_dashboard.iter_rows(min_row=ws_dashboard.min_row+1, max_row=ws_dashboard.max_row, min_col=1, max_col=1):
                for cell in row:
                    if (cell.value == item):
                        cell.value = ''
                        cell.offset(row=0, column=+1).value = ''
                        cell.offset(row=0, column=+2).value = ''
                        cell.offset(row=0, column=+3).value = ''
        except KeyError:
            print('NOT FOUND', item, name)
    print('TOTAL', found)
    print('AFTER', len(wb.sheetnames))
    wb.save('./Brokerage_ymetrics_report.xlsx')

def main(argv):
    clean()
    
if __name__ == "__main__":
    main(sys.argv[1:])