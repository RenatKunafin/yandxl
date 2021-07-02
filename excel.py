from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import Font, Fill, PatternFill
from datetime import datetime
from datetime import timedelta
from hashlib import md5
import time
import json
import re
import sys

from openpyxl.workbook.defined_name import ROW_RANGE_RE


class Excel:
    def __init__(self, cfg, data, startDate):
        self.data = data
        self.path_to_wb = cfg.get('smtp', 'PATH') + cfg.get('excel', 'WB_NAME')
        self.dashboard_ws_name = cfg.get('excel', 'DASHBOARD_WS_NAME')
        self.odbo_ws_name = cfg.get('excel', 'ODBO_FUNNEL_SHEET_NAME')
        self.odbo_iis_ws_name = cfg.get('excel', 'ODBO_IIS_FUNNEL_SHEET_NAME')
        self.odbo_titles = cfg.get('excel', 'ODBO_FUNNEL_TITLES').split(',')
        self.odbo_steps = cfg.get('excel', 'ODBO_FUNNEL_STEPS').split(',')
        self.titles_dashboard = cfg.get('excel', 'ROW_TITLES_DASHBOARD').split(',')
        self.date1 = startDate or cfg.get('yam', 'DATE1')
        self.title_font = Font(size=14, color='FF000000')
        self.odbo_process_name = 'Открытие брокерского счета'
        self.odbo_iis_process_name = 'Открытие брокерского счета и ИИС'

    def _build_dashboard_row(self, dimensions, metrics):
        entry = []
        for d in dimensions:
            entry.append(d['name'])
        return entry + metrics

    def _build_dashboard(self, wb):
        if not self.dashboard_ws_name in wb:
            ws_dashboard = wb.active
            ws_dashboard.title = self.dashboard_ws_name
            ws_dashboard.append(self.titles_dashboard)
            ws_dashboard.auto_filter.ref = ws_dashboard.dimensions
            for row in ws_dashboard.iter_rows(min_row=1, max_row=1, min_col=1, max_col=ws_dashboard.max_column):
                for cell in row:
                    cell.font = self.title_font
            ws_dashboard.column_dimensions['A'].width = 15
            ws_dashboard.column_dimensions['B'].width = 20
            ws_dashboard.column_dimensions['C'].width = 30
            ws_dashboard.column_dimensions['D'].width = 30
            ws_dashboard.column_dimensions['E'].width = 40
            ws_dashboard.column_dimensions['F'].width = 40
            ws_dashboard.column_dimensions['G'].width = 25
            ws_dashboard.column_dimensions['H'].width = 20
            ws_dashboard.column_dimensions['I'].width = 20
        ws_dashboard = wb[self.dashboard_ws_name]
        ws_dashboard.delete_rows(2, ws_dashboard.max_row)
        entries = []
        for d in self.data['data']:
            entries.append(self._build_dashboard_row(d['dimensions'], d['metrics']))
        for row in entries:
            ws_dashboard.append(row)

    def _dataToList(self, dimensions, metrics):
        return [dimensions[4]['name'], dimensions[0]['name'], dimensions[3]['name'], metrics[0]]

    def _sort(self, items):
        res = list()
        for pos in self.odbo_steps:
            for item in items:
                if pos == item[0]:
                    res.append(item)
        return res

    def _build_odbo_funnel(self, name):
        tree = {}
        for d in self.data['data']:
            # if (d['dimensions'][2]['name'] == name):
            process = d['dimensions'][2]['name']
            if process.endswith(name):
                step = d['dimensions'][4]['name']
                channel = d['dimensions'][0]['name']
                source = d['dimensions'][3]['name']
                action = d['dimensions'][5]['name']
                value = d['metrics'][0]
                key = step+'-'+channel+'-'+source
                if action == 'Открытие' or (step == 'Экран запрета' and d['dimensions'][6]['name'] == 'Открытие'):
                    tree[key] = [step, channel, source, value]
                elif (action == 'Один телефон' or action == 'Несколько телефонов') and d['dimensions'][6]['name'] == 'Открытие':
                    try:
                        tree[key][3] = tree[key][3] + value
                    except KeyError:
                        tree[key] = [step, channel, source, value]
        self._save_to('new_data.json', tree)
        temp = list(tree.values())
        res = list()
        for pos in self.odbo_steps:
            for item in temp:
                if pos == item[0]:
                    res.append(item)
        return res

    def _get_date(self):
        p = re.compile('(\d+)[A-z]+')
        res = p.findall(self.date1)
        if len(res) != 0:
            d = datetime.today() - timedelta(days=int(res[0]))
            return d.strftime('%d-%m-%Y')
        elif self.date1 is not None:
            return datetime.strptime(self.date1, '%Y-%m-%d').strftime('%d-%m-%Y')
        else:
            raise ValueError('Unknown date1 value', self.date1)

    def _save_to(self, file_name, data):
        f = open(file_name, "w")
        f.write(json.dumps(data))
        f.close()

    def _colect_old_data(self, ws):
        old_data = {}
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column, values_only=True):
            old_data[row[0]+'-'+row[1]+'-'+row[2]] = list(row)
        self._save_to('old_data.json', old_data)
        return old_data

    def _merge_data(self, ws, old_data, new_data):
        for item in new_data:
            key_from_new = item[0]+'-'+item[1]+'-'+item[2]
            value = item[3]
            if key_from_new in old_data:
                old_data[key_from_new].append(value)
            else:
                max_column = 4 if ws.max_column < 4 else ws.max_column
                res = [0] * (max_column - 3)
                res.append(value)
                temp = item[0:3] + res
                old_data[key_from_new] = temp
        result = self._sort(list(old_data.values()))
        self._save_to('result.json', result)
        return result

    def _update_ws(self, ws, data, date):
        ws.delete_rows(2, ws.max_row)
        # print('MAX', ws.max_column)
        max_column = 4 if ws.max_column < 4 else ws.max_column
        ws.cell(row = 1, column=max_column+1).value = date
        ws.cell(row = 1, column=max_column+1).font = self.title_font
        for row in data:
            ws.append(row)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=ws.max_column, max_col=ws.max_column):
            for cell in row:
                if cell.value is None:
                    cell.value = 0
        ws.column_dimensions[ws.cell(row = 1, column=ws.max_column).column_letter].width = 15

    def _log_file(self, wb, config, date):
        f = open('log_counter.txt', "r")
        counter = int(f.read() or 1)
        f.close()
        f = open('log_counter.txt', "w")
        counter = counter + 1
        
        print('COUNTER', counter)
        if counter >= 15:
            f.write('0')
            file_name = config.get('excel', 'WB_NAME')
            log_file_name = config.get('smtp', 'PATH') + file_name[0:-5] + '_' + date + '.xlsx'
            wb.save(log_file_name)
            print('REPORT LOGGED')
        else:
            f.write(str(counter))
        f.close()

    def _write_to_odbo_funnel(self, wb, date):
        ws = wb[self.odbo_ws_name]
        column = ws.max_column
        if column == 18:
            ws.delete_cols(4)
        new_data = self._build_odbo_funnel(self.odbo_process_name)
        old_data = self._colect_old_data(ws)
        result = self._merge_data(ws, old_data, new_data)
        self._update_ws(ws, result, date)

    def _write_to_odbo_iis_funnel(self, wb, date):
        ws = wb[self.odbo_iis_ws_name]
        column = ws.max_column
        if column == 18:
            ws.delete_cols(4)
        new_data = self._build_odbo_funnel(self.odbo_iis_process_name)
        old_data = self._colect_old_data(ws)
        result = self._merge_data(ws, old_data, new_data)
        self._update_ws(ws, result, date)

    def write_to_wb(self, config, mode=None):
        print('WRITE', self.path_to_wb)
        try:
            wb = load_workbook(self.path_to_wb)
            date = self._get_date()
            ws = wb[self.odbo_ws_name]
            self._log_file(wb, config, date)
            if ws.max_column == 18:
                wb = load_workbook(self.path_to_wb)
            self._build_dashboard(wb)
            self._write_to_odbo_funnel(wb, date)
            self._write_to_odbo_iis_funnel(wb, date)
            wb.save(self.path_to_wb)
            print('excel ready')
            
        except FileNotFoundError as e:
            print('Excel file not found', e)
            if mode == 'generation':
                self.init_wb()
            else:
                sys.exit(1)

    def _create_odbo_ws(self, wb, ws_name, process_name):
        ws_odbo = wb.create_sheet(ws_name)
        ws_odbo.title = ws_name
        ws_odbo.append(self.odbo_titles)
        for row in ws_odbo.iter_rows(min_row=1, max_row=1, min_col=1, max_col=ws_odbo.max_column):
                for cell in row:
                    cell.font = self.title_font
        ws_odbo.column_dimensions['A'].width = 20
        ws_odbo.column_dimensions['B'].width = 15
        ws_odbo.column_dimensions['C'].width = 30
        ws_odbo.column_dimensions['D'].width = 15
        ws_odbo.auto_filter.ref = ws_odbo.dimensions
        tree = self._build_odbo_funnel(process_name)
        for row in tree:
            ws_odbo.append(row)
        max_column = 4 if ws_odbo.max_column < 4 else ws_odbo.max_column
        ws_odbo.cell(row = 1, column=max_column).value = self._get_date()
        ws_odbo.cell(row = 1, column=max_column).font = self.title_font

    def init_wb(self):
        print('INIT')
        wb = Workbook()
        self._build_dashboard(wb)
        self._create_odbo_ws(wb, self.odbo_ws_name, self.odbo_process_name)
        self._create_odbo_ws(wb, self.odbo_iis_ws_name, self.odbo_iis_process_name)
        wb.save(self.path_to_wb)
        print('excel ready')
